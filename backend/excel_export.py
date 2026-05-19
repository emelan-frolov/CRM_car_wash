"""Модуль генерации Excel-файлов.

Вынесен в отдельный файл, чтобы при использовании ProcessPoolExecutor
дочерние процессы импортировали только этот лёгкий модуль, а не весь app.py
(где есть Flask, SQLAlchemy и подключение к БД).
"""
from io import BytesIO
from datetime import datetime


def generate_orders_excel(orders_data):
    """Генерация Excel-файла с заказами.
    
    Принимает только сериализуемые данные (list of dicts).
    Возвращает bytes готового xlsx-файла.
    Предназначена для запуска в отдельном процессе через ProcessPoolExecutor.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    headers = [
        'ID', 'Дата создания', 'Запланировано на', 'Завершен',
        'Клиент', 'Телефон', 'Автомобиль', 'Номер',
        'Услуги', 'Бокс', 'Сотрудник',
        'Длительность (мин)', 'Сумма (₽)', 'Оплачен', 'Статус', 'Примечания'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    status_map = {
        'pending': 'Ожидает',
        'in_progress': 'Выполняется',
        'completed': 'Завершен',
        'cancelled': 'Отменен'
    }
    
    def _fmt_dt(s):
        if not s:
            return ''
        try:
            return datetime.fromisoformat(s).strftime('%d.%m.%Y %H:%M')
        except (ValueError, TypeError):
            return s
    
    # Строки данных
    for row_idx, order in enumerate(orders_data, 2):
        row = [
            order['id'],
            _fmt_dt(order.get('created_at')),
            _fmt_dt(order.get('scheduled_time')),
            _fmt_dt(order.get('completed_time')),
            order.get('client_name') or '',
            order.get('client_phone') or '',
            order.get('car_info') or '',
            order.get('car_license_plate') or '',
            order.get('service_names') or '',
            order.get('box_name') or '',
            order.get('employee_name') or '',
            order.get('total_duration') or 0,
            order.get('total_price') or 0,
            'Да' if order.get('is_paid') else 'Нет',
            status_map.get(order.get('status'), order.get('status', '')),
            order.get('notes') or ''
        ]
        
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    
    # Ширина колонок
    column_widths = [6, 18, 18, 18, 25, 18, 22, 14, 35, 12, 22, 14, 12, 10, 14, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
